#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""TPEx 台灣債券成交歷史與債券主檔整合器。

功能：
1. 下載 BDdys01a（等殖 / EBTS）與 BDdcs001（處所 / OTC）日報表。
2. 維護 data/ebts_outright.json 與 data/otc_outright.json。
3. 下載 TPEx bond_ISSBD1_data 至 bond_ISSBD11_data 發行資料。
4. 依債券代號補上發行人、到期日、債券類型等主檔資訊。
5. 合併兩市場至 data/merged_outright.json。
6. 以本次執行的台北日期計算最新剩餘年期，四捨五入至小數第一位。

正式檔名固定為 update_bond_history.py。
"""
from __future__ import annotations

import argparse
import hashlib
import json
import logging
import math
import random
import re
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable
from zoneinfo import ZoneInfo

import requests
import xlrd

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
DEBUG_DIR = ROOT / "debug"
TAIPEI = ZoneInfo("Asia/Taipei")
REPORT_BASE_URL = "https://www.tpex.org.tw/storage/bond_zone/tradeinfo/govbond"
OPENAPI_BASE_URL = "https://www.tpex.org.tw/openapi/v1"
CODE_RE = re.compile(r"^[A-Z0-9][A-Z0-9]{4,11}$", re.I)
DATE_DIGITS_RE = re.compile(r"\D")

REPORTS = {
    "BDdys01a": {
        "market": "EBTS",
        "market_zh": "等殖",
        "output": DATA_DIR / "ebts_outright.json",
        "description": "等殖成交行情表（買賣斷）",
    },
    "BDdcs001": {
        "market": "OTC",
        "market_zh": "處所",
        "output": DATA_DIR / "otc_outright.json",
        "description": "營業處所成交行情表（買賣斷）",
    },
}

BOND_APIS = {
    "bond_ISSBD1_data": "政府公債",
    "bond_ISSBD2_data": "外國金融債",
    "bond_ISSBD3_data": "金融債券",
    "bond_ISSBD4_data": "普通公司債",
    "bond_ISSBD5_data": "轉換或交換公司債",
    "bond_ISSBD6_data": "海外轉換公司債",
    "bond_ISSBD7_data": "附認股權公司債",
    "bond_ISSBD8_data": "海外附認股權公司債",
    "bond_ISSBD9_data": "海外普通債",
    "bond_ISSBD10_data": "國際債券",
    "bond_ISSBD11_data": "國際債券外國發行人",
}

FIELD_ALIASES = {
    "bond_code": ["債券代碼", "債券代號", "BondCode", "Bond Code", "Code"],
    "bond_name": ["債券簡稱", "債券名稱", "BondName", "ShortName", "Name"],
    "issuer_code": ["機構代碼", "發行公司代號", "發行人代號", "IssuerCode"],
    "issuer_name": ["機構名稱", "發行公司名稱", "發行人名稱", "發行人", "IssuerName", "Issuer"],
    "issue_date": ["發行日期", "發行日", "IssueDate", "Issuing Date"],
    "maturity_date": ["到期日期", "到期日", "MaturityDate", "Maturity Date"],
    "coupon_rate": ["票面利率", "利率", "CouponRate", "Coupon (%)", "Coupon"],
    "currency": ["幣別", "Currency"],
    "issue_amount": ["發行總額", "發行金額", "發行面額", "IssueAmount"],
    "outstanding_amount": ["目前餘額", "流通餘額", "OutstandingAmount"],
}


def setup_logging() -> None:
    DEBUG_DIR.mkdir(exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(DEBUG_DIR / "update.log", encoding="utf-8"),
        ],
    )


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float):
        if math.isnan(value):
            return None
        return int(value) if value.is_integer() else value
    text = str(value).replace("\u3000", " ").strip()
    return text or None


def normalize_key(value: Any) -> str:
    return re.sub(r"[\s_()（）%％/\-]", "", str(value or "")).lower()


def get_alias(row: dict[str, Any], alias_name: str) -> Any:
    normalized = {normalize_key(k): v for k, v in row.items()}
    for alias in FIELD_ALIASES[alias_name]:
        value = normalized.get(normalize_key(alias))
        if value not in (None, ""):
            return clean_value(value)
    return None


def normalize_bond_code(value: Any) -> str | None:
    text = str(value or "").strip().upper().replace(" ", "")
    return text if CODE_RE.fullmatch(text) else None


def base_bond_code(code: str) -> str:
    # 政府增額公債末碼 R 與原券共用到期日。精確代號仍優先查詢。
    return code[:-1] if code.startswith("A") and code.endswith("R") else code


def parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if any(word in text for word in ("無到期", "永續", "N/A", "NA")):
        return None
    digits = DATE_DIGITS_RE.sub("", text)
    try:
        if len(digits) == 8:
            year = int(digits[:4])
            return date(year, int(digits[4:6]), int(digits[6:8]))
        if len(digits) == 7:
            year = int(digits[:3]) + 1911
            return date(year, int(digits[3:5]), int(digits[5:7]))
    except ValueError:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def format_date(value: Any) -> str | None:
    parsed = parse_date(value)
    return parsed.isoformat() if parsed else None


def remaining_years(maturity: str | None, as_of: date) -> float | None:
    maturity_day = parse_date(maturity)
    if maturity_day is None:
        return None
    return round((maturity_day - as_of).days / 365.25, 1)


def maturity_status(maturity: str | None, is_perpetual: bool, as_of: date) -> str:
    if is_perpetual:
        return "perpetual"
    maturity_day = parse_date(maturity)
    if maturity_day is None:
        return "unknown"
    return "matured" if maturity_day < as_of else "outstanding"


def new_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/126 Safari/537.36",
        "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.7",
        "Referer": "https://www.tpex.org.tw/",
    })
    try:
        session.get("https://www.tpex.org.tw/", timeout=30)
    except requests.RequestException:
        pass
    return session


def request_bytes(session: requests.Session, url: str, accept: str, attempts: int = 3) -> bytes | None:
    for attempt in range(1, attempts + 1):
        try:
            response = session.get(url, timeout=60, headers={"Accept": accept})
            logging.info("GET %s -> %s bytes=%s attempt=%s", url, response.status_code, len(response.content), attempt)
            if response.status_code == 404:
                return None
            if response.status_code in (403, 429, 500, 502, 503, 504):
                time.sleep((2 ** attempt) + random.random())
                continue
            response.raise_for_status()
            return response.content
        except requests.RequestException as exc:
            logging.warning("下載失敗 %s: %s", url, exc)
            time.sleep((2 ** attempt) + random.random())
    return None


def report_url(report: str, day: date) -> str:
    ymd = day.strftime("%Y%m%d")
    return f"{REPORT_BASE_URL}/{day:%Y}/{day:%Y%m}/{report}.{ymd}-C.xls"


def unique_headers(values: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(values, 1):
        base = str(clean_value(value) or f"column_{index}").replace("\n", " ").strip()
        seen[base] = seen.get(base, 0) + 1
        headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return headers


def parse_xls(content: bytes, report: str, day: date, source_url: str) -> dict[str, Any]:
    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    matrix = [[clean_value(sheet.cell_value(r, c)) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
    header_row = None
    for index, row in enumerate(matrix[:35]):
        joined = " ".join(str(item or "") for item in row)
        if "代號" in joined or "債券代碼" in joined or "Code" in joined:
            header_row = index
    if header_row is None:
        raise ValueError(f"找不到表頭: {report} {day}")
    headers = unique_headers(matrix[header_row])
    records: list[dict[str, Any]] = []
    market = REPORTS[report]["market"]
    market_zh = REPORTS[report]["market_zh"]
    for row in matrix[header_row + 1:]:
        values = row[:len(headers)] + [None] * max(0, len(headers) - len(row))
        candidates = [str(v).strip() for v in values[:5] if v not in (None, "")]
        if not candidates:
            continue
        if any(value.lower() in ("合計", "total", "註", "remark") for value in candidates):
            break
        code = next((normalize_bond_code(value) for value in candidates if normalize_bond_code(value)), None)
        if not code:
            continue
        item = {header: value for header, value in zip(headers, values) if value is not None}
        item.update({"bond_code": code, "market": market, "market_zh": market_zh})
        records.append(item)
    return {
        "date": day.isoformat(),
        "report": report,
        "market": market,
        "market_zh": market_zh,
        "source_url": source_url,
        "sha256": hashlib.sha256(content).hexdigest(),
        "sheet_name": sheet.name,
        "header_row_1_based": header_row + 1,
        "columns": headers,
        "record_count": len(records),
        "records": records,
    }


def load_store(path: Path, report: str, description: str) -> dict[str, Any]:
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return {"metadata": {"dataset": report, "description": description, "last_updated": None}, "dates": {}}


def atomic_json_write(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    temp.replace(path)


def save_store(path: Path, store: dict[str, Any]) -> None:
    store["metadata"]["last_updated"] = datetime.now(TAIPEI).isoformat(timespec="seconds")
    store["dates"] = dict(sorted(store["dates"].items()))
    atomic_json_write(path, store)


def fetch_api_rows(session: requests.Session, endpoint: str) -> list[dict[str, Any]]:
    url = f"{OPENAPI_BASE_URL}/{endpoint}"
    content = request_bytes(session, url, "application/json,text/csv;q=0.8,*/*;q=0.5")
    if content is None:
        raise RuntimeError(f"API無回應: {endpoint}")
    payload = json.loads(content.decode("utf-8-sig"))
    if isinstance(payload, dict):
        for key in ("data", "aaData", "records", "result"):
            if isinstance(payload.get(key), list):
                payload = payload[key]
                break
    if not isinstance(payload, list):
        raise ValueError(f"API格式不是陣列: {endpoint}")
    return [row for row in payload if isinstance(row, dict)]


def make_master_record(row: dict[str, Any], endpoint: str, category: str) -> dict[str, Any] | None:
    code = normalize_bond_code(get_alias(row, "bond_code"))
    if not code:
        return None
    maturity_raw = get_alias(row, "maturity_date")
    maturity = format_date(maturity_raw)
    maturity_text = str(maturity_raw or "")
    is_perpetual = maturity is None and any(word in maturity_text for word in ("無到期", "永續"))
    return {
        "bond_code": code,
        "base_bond_code": base_bond_code(code),
        "bond_name": get_alias(row, "bond_name"),
        "issuer_code": get_alias(row, "issuer_code"),
        "issuer_name": get_alias(row, "issuer_name"),
        "bond_type": category,
        "issue_date": format_date(get_alias(row, "issue_date")),
        "maturity_date": maturity,
        "coupon_rate": get_alias(row, "coupon_rate"),
        "currency": get_alias(row, "currency"),
        "issue_amount": get_alias(row, "issue_amount"),
        "outstanding_amount": get_alias(row, "outstanding_amount"),
        "is_perpetual": is_perpetual,
        "source_endpoint": endpoint,
        "raw": row,
    }


def fetch_bond_master(session: requests.Session, as_of: date) -> dict[str, Any]:
    old_path = DATA_DIR / "bond_master.json"
    old_payload: dict[str, Any] = {}
    if old_path.exists():
        old_payload = json.loads(old_path.read_text(encoding="utf-8"))
    old_records = old_payload.get("bonds", {})

    records: dict[str, dict[str, Any]] = {}
    source_status: dict[str, Any] = {}
    successful_endpoints: set[str] = set()

    for endpoint, category in BOND_APIS.items():
        try:
            rows = fetch_api_rows(session, endpoint)
            accepted = 0
            for row in rows:
                record = make_master_record(row, endpoint, category)
                if record is None:
                    continue
                record["latest_remaining_years"] = remaining_years(record["maturity_date"], as_of)
                record["maturity_status"] = maturity_status(
                    record["maturity_date"], record["is_perpetual"], as_of
                )
                records[record["bond_code"]] = record
                accepted += 1
            successful_endpoints.add(endpoint)
            source_status[endpoint] = {"ok": True, "rows": len(rows), "accepted": accepted}
        except Exception as exc:
            logging.exception("債券主檔API失敗: %s", endpoint)
            source_status[endpoint] = {"ok": False, "error": str(exc), "fallback": "existing bond_master.json"}

    # 單一API暫時失敗時，只沿用該API舊資料，不讓其他成功來源把歷史主檔刪掉。
    for code, old_record in old_records.items():
        endpoint = old_record.get("source_endpoint")
        if endpoint not in successful_endpoints and code not in records:
            preserved = dict(old_record)
            preserved["latest_remaining_years"] = remaining_years(preserved.get("maturity_date"), as_of)
            preserved["maturity_status"] = maturity_status(
                preserved.get("maturity_date"), bool(preserved.get("is_perpetual")), as_of
            )
            records[code] = preserved

    if not records:
        raise RuntimeError("所有債券主檔API均失敗，且沒有既有 bond_master.json 可沿用")

    payload = {
        "metadata": {
            "description": "TPEx 各類債券發行資料主檔，保留標準化欄位與各API原始raw資料",
            "as_of_date": as_of.isoformat(),
            "updated_at": datetime.now(TAIPEI).isoformat(timespec="seconds"),
            "record_count": len(records),
            "source_status": source_status,
        },
        "bonds": dict(sorted(records.items())),
    }
    atomic_json_write(DATA_DIR / "bond_master.json", payload)
    return payload


def find_master(code: str, bonds: dict[str, dict[str, Any]]) -> tuple[dict[str, Any] | None, str | None]:
    if code in bonds:
        return bonds[code], code
    base = base_bond_code(code)
    if base in bonds:
        return bonds[base], base
    # 主檔偶爾只留增額券或原券其中之一，最後再比對 base_bond_code。
    for master_code, record in bonds.items():
        if record.get("base_bond_code") == base:
            return record, master_code
    return None, None


def iter_market_records(store: dict[str, Any]) -> Iterable[tuple[str, dict[str, Any]]]:
    for trade_date, day_payload in store.get("dates", {}).items():
        for record in day_payload.get("records", []):
            yield trade_date, record


def build_merged(as_of: date, master_payload: dict[str, Any]) -> dict[str, Any]:
    bonds = master_payload.get("bonds", {})
    merged_dates: dict[str, list[dict[str, Any]]] = {}
    market_counts = {"EBTS": 0, "OTC": 0}
    unmapped: set[str] = set()
    for report, cfg in REPORTS.items():
        store = load_store(cfg["output"], report, cfg["description"])
        for trade_date, source_record in iter_market_records(store):
            code = normalize_bond_code(source_record.get("bond_code"))
            if not code:
                continue
            master, matched_code = find_master(code, bonds)
            if master is None:
                unmapped.add(code)
            merged_record = dict(source_record)
            merged_record.update({
                "trade_date": trade_date,
                "market": cfg["market"],
                "market_zh": cfg["market_zh"],
                "source_report": report,
                "bond_code": code,
                "base_bond_code": base_bond_code(code),
                "master_match_code": matched_code,
                "master_mapped": master is not None,
                "bond_name_master": master.get("bond_name") if master else None,
                "issuer_code": master.get("issuer_code") if master else None,
                "issuer_name": master.get("issuer_name") if master else None,
                "bond_type": master.get("bond_type") if master else None,
                "issue_date": master.get("issue_date") if master else None,
                "maturity_date": master.get("maturity_date") if master else None,
                "coupon_rate": master.get("coupon_rate") if master else None,
                "currency": master.get("currency") if master else None,
                "is_perpetual": master.get("is_perpetual", False) if master else False,
                "latest_remaining_years": remaining_years(master.get("maturity_date"), as_of) if master else None,
                "remaining_years_as_of": as_of.isoformat(),
            })
            merged_dates.setdefault(trade_date, []).append(merged_record)
            market_counts[cfg["market"]] += 1
    for records in merged_dates.values():
        records.sort(key=lambda item: (item.get("bond_code", ""), item.get("market", "")))
    output = {
        "metadata": {
            "description": "等殖與處所買賣斷成交歷史合併資料",
            "updated_at": datetime.now(TAIPEI).isoformat(timespec="seconds"),
            "remaining_years_as_of": as_of.isoformat(),
            "remaining_years_formula": "round((maturity_date - as_of_date).days / 365.25, 1)",
            "date_count": len(merged_dates),
            "record_count": sum(len(rows) for rows in merged_dates.values()),
            "market_record_counts": market_counts,
            "unmapped_bond_count": len(unmapped),
            "unmapped_bond_codes": sorted(unmapped),
        },
        "dates": dict(sorted(merged_dates.items())),
    }
    atomic_json_write(DATA_DIR / "merged_outright.json", output)
    return output



# =============================================================================
# OPTIONAL EXCEL EXPORT FUNCTION
# 若未來不需要 Excel，只要刪除這整段 export_merged_to_excel() 即可。
# main() 使用 globals().get() 動態尋找函式，刪除後不會影響 JSON 更新流程。
# =============================================================================
def export_merged_to_excel(
    merged_payload: dict[str, Any],
    output_path: Path | None = None,
) -> Path:
    """將 merged_outright 的所有逐筆成交資料輸出為 Excel。

    工作表：
    - Transactions：每一筆 EBTS / OTC 成交各占一列。
    - Metadata：合併檔 metadata，方便核對產製日期與未 Mapping 券號。

    動態原始欄位會完整保留。dict/list 會序列化為 JSON 字串，避免資料遺失。
    """
    # 放在函式內匯入，刪除整個函式後，主程式不再依賴 openpyxl。
    # 若 requirements.txt 尚未加入 openpyxl，GitHub Actions 會在此自動安裝。
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        import subprocess
        import sys

        logging.warning("偵測不到 openpyxl，開始自動安裝 openpyxl==3.1.5")
        subprocess.check_call([
            sys.executable,
            "-m",
            "pip",
            "install",
            "--disable-pip-version-check",
            "openpyxl==3.1.5",
        ])
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

    output_path = output_path or (DATA_DIR / "merged_outright.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows: list[dict[str, Any]] = []
    for trade_date, records in merged_payload.get("dates", {}).items():
        for record in records:
            item = dict(record)
            item.setdefault("trade_date", trade_date)
            rows.append(item)

    preferred_columns = [
        "trade_date",
        "market",
        "market_zh",
        "source_report",
        "bond_code",
        "base_bond_code",
        "master_match_code",
        "master_mapped",
        "bond_name_master",
        "issuer_code",
        "issuer_name",
        "bond_type",
        "issue_date",
        "maturity_date",
        "latest_remaining_years",
        "remaining_years_as_of",
        "coupon_rate",
        "currency",
        "is_perpetual",
    ]
    all_columns = {key for row in rows for key in row.keys()}
    columns = [column for column in preferred_columns if column in all_columns]
    columns.extend(sorted(all_columns - set(columns)))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transactions"
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_index, column_name in enumerate(columns, 1):
        cell = sheet.cell(row=1, column=column_index, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    def excel_value(value: Any) -> Any:
        if isinstance(value, (dict, list, tuple, set)):
            return json.dumps(value, ensure_ascii=False, sort_keys=True)
        if value is None:
            return None
        if isinstance(value, (str, int, float, bool, datetime, date)):
            return value
        return str(value)

    for row_index, record in enumerate(rows, 2):
        for column_index, column_name in enumerate(columns, 1):
            value = excel_value(record.get(column_name))
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(vertical="top")
            if column_name == "latest_remaining_years" and isinstance(value, (int, float)):
                cell.number_format = "0.0"

    if columns:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, len(rows) + 1)}"

    # 以抽樣內容估算欄寬，避免逐格掃描大型歷史檔造成執行時間過長。
    sample_rows = rows[:1000]
    for column_index, column_name in enumerate(columns, 1):
        lengths = [len(str(column_name))]
        for record in sample_rows:
            value = excel_value(record.get(column_name))
            if value is not None:
                lengths.append(len(str(value)))
        width = min(max(max(lengths) + 2, 10), 42)
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    metadata_sheet = workbook.create_sheet("Metadata")
    metadata_sheet.sheet_view.showGridLines = False
    metadata_sheet.append(["key", "value"])
    for cell in metadata_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    for key, value in merged_payload.get("metadata", {}).items():
        metadata_sheet.append([key, excel_value(value)])
    metadata_sheet.column_dimensions["A"].width = 30
    metadata_sheet.column_dimensions["B"].width = 100
    for row in metadata_sheet.iter_rows(min_row=2, max_col=2):
        row[1].alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(output_path)
    logging.info("Excel逐筆明細已輸出: %s rows=%s columns=%s", output_path, len(rows), len(columns))
    return output_path


def daterange(start: date, end: date) -> Iterable[date]:
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def update_trade_history(session: requests.Session, days: Iterable[date], force: bool, sleep_seconds: float) -> int:
    changed = 0
    day_list = list(days)
    for report, cfg in REPORTS.items():
        store = load_store(cfg["output"], report, cfg["description"])
        for day in day_list:
            key = day.isoformat()
            if key in store.get("dates", {}) and not force:
                continue
            url = report_url(report, day)
            content = request_bytes(session, url, "application/vnd.ms-excel,application/octet-stream;q=0.9,*/*;q=0.5")
            if content is None:
                continue
            if len(content) < 512 or content[:4] not in (b"\xd0\xcf\x11\xe0", b"PK\x03\x04"):
                logging.warning("不是有效Excel: %s", url)
                continue
            try:
                store.setdefault("dates", {})[key] = parse_xls(content, report, day, url)
                save_store(cfg["output"], store)
                changed += 1
            except Exception:
                bad_path = DEBUG_DIR / f"{report}.{day:%Y%m%d}.xls"
                bad_path.write_bytes(content)
                logging.exception("解析失敗，原檔已保存: %s", bad_path)
            time.sleep(sleep_seconds)
    return changed


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=("bootstrap", "daily", "rebuild"), default="daily")
    parser.add_argument("--start", default="2026-01-01")
    parser.add_argument("--end")
    parser.add_argument("--date")
    parser.add_argument("--as-of", help="剩餘年期計算基準日，預設台北當日")
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--sleep", type=float, default=0.35)
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="不輸出 data/merged_outright.xlsx；預設會輸出逐筆明細Excel",
    )
    args = parser.parse_args()
    setup_logging()

    today = datetime.now(TAIPEI).date()
    as_of = date.fromisoformat(args.as_of) if args.as_of else today
    if args.mode == "daily":
        days = [date.fromisoformat(args.date) if args.date else today]
    elif args.mode == "bootstrap":
        start = date.fromisoformat(args.start)
        end = date.fromisoformat(args.end) if args.end else today
        days = list(daterange(start, end))
    else:
        days = []

    session = new_session()
    changed = update_trade_history(session, days, args.force, args.sleep) if days else 0
    master = fetch_bond_master(session, as_of)
    merged = build_merged(as_of, master)

    # 使用動態查找，未來若刪除 OPTIONAL EXCEL EXPORT FUNCTION，不會影響JSON流程。
    excel_exporter = globals().get("export_merged_to_excel")
    if not args.no_excel and callable(excel_exporter):
        excel_exporter(merged, DATA_DIR / "merged_outright.xlsx")

    logging.info(
        "完成 trade_changed=%s master=%s merged=%s unmapped=%s",
        changed,
        master["metadata"]["record_count"],
        merged["metadata"]["record_count"],
        merged["metadata"]["unmapped_bond_count"],
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
